import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

def procurar(primeiro_objeto, segundo_objeto, primeiro_obj, segundo_obj, caminho_arquivo1, global_workbook, colunas_desejadas):
    # Carregar dados do Excel
    df = pd.read_excel(caminho_arquivo1, header=6)
    
    # Converta os valores na coluna 'DAY' para strings
    df['DAY'] = df['DAY'].astype(str)
    # Remover ".0" apenas dos valores que terminam com ".0"
    df['DAY'] = df['DAY'].apply(lambda x: x[:-2] if x.endswith('.0') else x)
    # Mapear a conversão para inteiros apenas para os valores numéricos
    df['DAY'] = df['DAY'].map(lambda x: int(x) if x.isdigit() else x)

    # Objeto desejado
    objeto_desejado = f'{primeiro_objeto}{segundo_objeto}'
    # Lista de eventos desejados
    eventos_desejados = [f'OBJ{primeiro_obj}', f'OBJ{segundo_obj}']

    ws = global_workbook.create_sheet(title=f'{primeiro_objeto}_{segundo_objeto}')

    # Inicializa uma lista para armazenar os DataFrames finais de cada evento
    dfs_finais = []

    # Inicializa df_primeiro_objeto como None
    df_primeiro_objeto = None

    # Adicione a verificação aqui
    if df_primeiro_objeto is None:
        df_primeiro_objeto = pd.DataFrame(columns=colunas_desejadas)

    for evento_desejado in eventos_desejados:
        # Constrói o nome do evento
        evento = f'Mouse 1 sniffing On {evento_desejado}'
        # Filtra os dados pelo evento e objeto desejado
        dados_do_evento = df[(df['Events'] == evento) & (df['OBJECTS'].str.contains(objeto_desejado))]
        # Define as colunas desejadas
        colunas_desejadas = ['DAY', 'ANIMAL', 'OBJECTS', 'Total Bouts', 'Total Duration(Second)', 'Latency(Second)', 'Ending time(Second) of First Bout']
        # Inicializa uma lista vazia para armazenar os dados filtrados
        dados_filtrados = []
        # Inicializa a variável para armazenar o último dia processado
        ultimo_dia = None

        # Itera sobre as linhas dos dados do evento desejado
        for _, row in dados_do_evento.iterrows():
            # Adiciona uma linha em branco antes do início de cada novo conjunto de dias
            if ultimo_dia is not None and row['DAY'] != ultimo_dia:
                dados_filtrados.append([None] * len(colunas_desejadas))

            # Inicializa uma lista para armazenar os dados desta linha
            linha = []

            # Adiciona os dados da linha à lista de linhas filtradas
            for coluna in colunas_desejadas:
                # Adiciona o valor da coluna atual
                linha.append(row[coluna])
                # Adiciona uma coluna em branco à direita, exceto para 'DAY', 'ANIMAL' e 'OBJECTS'
                if coluna not in ['DAY', 'ANIMAL', 'OBJECTS']:
                    linha.append(None)

            # Adiciona a linha completa aos dados filtrados
            dados_filtrados.append(linha)
            ultimo_dia = row['DAY']

        # Cria um DataFrame a partir dos dados filtrados
        colunas_finais = []
        for col in colunas_desejadas:
            colunas_finais.append(col)
            if col not in ['DAY', 'ANIMAL', 'OBJECTS']:
                colunas_finais.append(f"{col}_{segundo_objeto}")
        df_final = pd.DataFrame(dados_filtrados, columns=colunas_finais)


        # Se for o evento do primeiro objeto, armazena os dados no DataFrame correspondente
        if evento_desejado == f'OBJ{primeiro_obj}':
            df_primeiro_objeto = df_final
        # Se for o evento do segundo objeto, adiciona os dados como colunas novas no DataFrame do primeiro objeto
        else:
            # Itera sobre as linhas dos dados do evento desejado
            for _, row in df_final.iterrows():
                # Encontra os índices das linhas correspondentes no DataFrame do primeiro objeto
                idx = df_primeiro_objeto[(df_primeiro_objeto['DAY'] == row['DAY']) & (df_primeiro_objeto['ANIMAL'] == row['ANIMAL'])].index

                # Itera sobre os índices encontrados
                for i in idx:
                    # Adiciona os dados do segundo objeto nas células em branco correspondentes
                    for col in colunas_desejadas[3:]:  # Ignora 'DAY', 'ANIMAL', 'OBJECTS'
                        df_primeiro_objeto.at[i, f"{col}_{segundo_objeto}"] = row[col]

                    # Encontra o índice da coluna Total Duration(Second)_{segundo_objeto}
                    idx_total_duration = df_primeiro_objeto.columns.get_loc(f"Total Duration(Second)_{segundo_objeto}")

                    # Verifica se as colunas "Blank Column 1" e "Blank Column 2" já existem
                    if 'Blank Column 1' not in df_primeiro_objeto.columns:
                        df_primeiro_objeto.insert(idx_total_duration + 1, 'Blank Column 1', None)
                    if 'Blank Column 2' not in df_primeiro_objeto.columns:
                        df_primeiro_objeto.insert(idx_total_duration + 2, 'Blank Column 2', None)

                    # Realiza os cálculos para as colunas adicionadas
                    subtrair = df_primeiro_objeto.at[i, f"Total Duration(Second)_{segundo_objeto}"] - df_primeiro_objeto.at[i, "Total Duration(Second)"]
                    soma = df_primeiro_objeto.at[i, "Total Duration(Second)"] + df_primeiro_objeto.at[i, f"Total Duration(Second)_{segundo_objeto}"]
                    if pd.notna(subtrair) and pd.notna(soma) and soma != 0:
                        df_primeiro_objeto.at[i, 'Blank Column 1'] = soma
                        df_primeiro_objeto.at[i, 'Blank Column 2'] = round(subtrair / soma, 2)
                    else:
                        # Se algum dos valores não for numérico ou for NaN, atribui NaN às colunas 'Blank Column 1' e 'Blank Column 2'
                        df_primeiro_objeto.at[i, 'Blank Column 1'] = float('nan')
                        df_primeiro_objeto.at[i, 'Blank Column 2'] = float('nan')

    # Adiciona o DataFrame final do primeiro objeto à lista de DataFrames finais
    dfs_finais.append(df_primeiro_objeto)
    # Concatena os DataFrames finais ao longo do eixo das colunas
    df_final_concatenado = pd.concat(dfs_finais, axis=1)

    # Preenche valores NA ou NaN na coluna 'Events' com uma string vazia
    df['Events'] = df['Events'].fillna('')

    # Inicializa uma lista para armazenar os valores da coluna "Droga"
    valores_droga = []

    # Itera sobre as linhas do DataFrame final
    for _, row in df_final_concatenado.iterrows():
        # Obtém o evento correspondente ao primeiro objeto
        evento_primeiro_objeto = f'Mouse 1 sniffing On OBJ{primeiro_obj}'
        # Filtra o DataFrame original para encontrar a linha correspondente ao evento do primeiro objeto nesta linha
        linha_correspondente = df[(df['Events'] == evento_primeiro_objeto) & (df['DAY'] == row['DAY']) & (df['ANIMAL'] == row['ANIMAL'])]
        # Se encontrar uma linha correspondente, adiciona o valor da coluna "DRUG" à lista de valores de droga
        if not linha_correspondente.empty:
            valores_droga.append(linha_correspondente['DRUG'].iloc[0])
        else:
            # Se não encontrar uma linha correspondente, adiciona um valor vazio
            valores_droga.append('')

    # Adiciona os valores da coluna "Droga" ao DataFrame final
    df_final_concatenado['Droga'] = valores_droga

    for row in dataframe_to_rows(df_final_concatenado, index=False, header=True):
        ws.append(row)

    # Define as colunas a serem mescladas e seus respectivos pares do segundo objeto
    colunas_a_mesclar = ['D', 'F', 'J', 'L']
    pares_segundo_objeto = ['E', 'G', 'K', 'M']
    coluna_separada = ['H', 'I']

    # Itera sobre as colunas a serem mescladas
    for coluna, par_segundo_objeto in zip(colunas_a_mesclar, pares_segundo_objeto):
        ws.merge_cells(f'{coluna}1:{par_segundo_objeto}1')

    # Define o tamanho desejado para as colunas separadas
    for col_separada in coluna_separada:
        ws.column_dimensions[col_separada].width = 16

    # Define o tamanho desejado para as colunas a serem mescladas e seus pares
    for col in colunas_a_mesclar + pares_segundo_objeto:
        ws.column_dimensions[col].width = 12

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column):
        for cell in row:
            # Centraliza o conteúdo da célula
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            # Verifica se a célula está preenchida
            if cell.value is not None:
                # Centraliza o conteúdo da célula
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Renomeia as colunas
    novo_nome_colunas = {
        'OBJECTS': 'Objetos',
        'DAY': 'Dia',
        'ANIMAL': 'Animal',
        'Total Bouts': 'Bouts',
        'Total Duration(Second)': 'Exploração',
        'Blank Column 1': 'Total',
        'Blank Column 2': 'DI',
        'Latency(Second)': 'Latência',
        'Ending time(Second) of First Bout': 'FIM'
    }

    # Itera sobre as células de cabeçalho para renomear as colunas
    for coluna_antiga, novo_nome in novo_nome_colunas.items():
        # Obtém o número da coluna antiga
        for cell in ws[1]:
            if cell.value == coluna_antiga:
                numero_coluna_antiga = cell.column
                break
        # Define o novo nome da coluna
        ws.cell(row=1, column=numero_coluna_antiga, value=novo_nome)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

    return objeto_desejado, eventos_desejados  # Retorna os resultados ao final da função
