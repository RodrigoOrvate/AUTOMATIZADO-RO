import tkinter as tk
from tkinter import filedialog
import pandas as pd
import openpyxl
import subprocess
import os
import sys
from functools import partial
from PIL import Image, ImageTk

# Importando as funções modificadas
from procurar_objeto import procurar
from procurar_distvel import organizar

# Configurações Globais
caminho_arquivo1 = ""
caminho_arquivo2 = ""
conjuntos_objetos = []
global_workbook = openpyxl.Workbook()
global_excel_filename_obj = "dados_filtrados_obj.xlsx"
global_excel_filename_distvel = "dados_filtrados_distvel.xlsx"
erro_exibido = False
colunas_desejadas = ['DAY', 'ANIMAL', 'OBJECTS', 'Total Bouts', 'Total Duration(Second)', 'Latency(Second)', 'Ending time(Second) of First Bout']
entry_label_style = {"bg": "#F0F0F0", "fg": "#000000", "font": ("Helvetica", 10), "bd": 0, "highlightthickness": 2, "highlightbackground": "#4CAF50", "highlightcolor": "#4CAF50", "relief": "flat"}
pares_objetos = set()
objs = set()

# --- Funções do Sistema (Mantidas ou Levemente Ajustadas) ---

def pesquisar_arquivo1():
    global caminho_arquivo1
    filename1 = filedialog.askopenfilename()
    caminho_arquivo1 = filename1
    caminho_entry1.config(state='normal')
    caminho_entry1.delete(0, tk.END) # Garante que limpa antes de inserir
    caminho_entry1.insert(0, filename1)
    caminho_entry1.config(state='readonly')
    liberar_criar_conjuntos()
    atualizar_rotulos() # Chama atualização ao carregar

def pesquisar_arquivo2():
    global caminho_arquivo2
    filename2 = filedialog.askopenfilename()
    caminho_arquivo2 = filename2
    caminho_entry2.config(state='normal')
    caminho_entry2.delete(0, tk.END)
    caminho_entry2.insert(0, filename2)
    caminho_entry2.config(state='readonly')
    liberar_botaodois()

def limpar_entry1():
    global caminho_arquivo1
    caminho_entry1.config(state='normal')  
    caminho_entry1.delete(0, tk.END)  
    caminho_entry1.config(state='readonly')
    caminho_arquivo1 = "" # Limpa a variável
    quantidade_entry.delete(0, tk.END)
    # Limpa labels de objetos
    pares_objetos_var.set("")
    objs_var.set("")

def limpar_entry2():
    global caminho_arquivo2
    caminho_entry2.config(state='normal')  
    caminho_entry2.delete(0, tk.END)  
    caminho_entry2.config(state='readonly')
    caminho_arquivo2 = "" # Limpa a variável

# ... (Funções criar_conjunto_labels, on_entry_click, on_focusout mantidas iguais) ...

def criar_conjunto_labels(conjuntos_frame, numero_conjunto):
    conjunto_frame = tk.Frame(conjuntos_frame, relief="solid", bd=1)
    conjunto_frame.grid(row=numero_conjunto-1, column=0, padx=5, pady=5, sticky="ew")
    
    titulo_label = tk.Label(conjunto_frame, text=f"Planilha {numero_conjunto}", font=("Arial", 12, "bold"))
    titulo_label.grid(row=0, column=0, columnspan=4, pady=(10, 5))

    objeto1_entry = tk.Entry(conjunto_frame, font=("Helvetica", 10, "bold"))
    objeto1_entry.insert(0, "Primeiro par_objeto")
    objeto1_entry.config(fg="grey")
    objeto1_entry.bind("<FocusIn>", lambda event: on_entry_click(objeto1_entry, "Primeiro par_objeto"))
    objeto1_entry.bind("<FocusOut>", lambda event: on_focusout(objeto1_entry, "Primeiro par_objeto"))
    objeto1_entry.grid(row=1, column=1, padx=(0, 5), pady=5, sticky="ew")

    objeto2_entry = tk.Entry(conjunto_frame, font=("Helvetica", 10, "bold"))
    objeto2_entry.insert(0, "Segundo par_objeto")
    objeto2_entry.config(fg="grey")
    objeto2_entry.bind("<FocusIn>", lambda event: on_entry_click(objeto2_entry, "Segundo par_objeto"))
    objeto2_entry.bind("<FocusOut>", lambda event: on_focusout(objeto2_entry, "Segundo par_objeto"))
    objeto2_entry.grid(row=1, column=3, padx=(0, 5), pady=5, sticky="ew")

    obj1_entry = tk.Entry(conjunto_frame, font=("Helvetica", 10, "bold"))
    obj1_entry.insert(0, "Primeiro OBJ")
    obj1_entry.config(fg="grey")
    obj1_entry.bind("<FocusIn>", lambda event: on_entry_click(obj1_entry, "Primeiro OBJ"))
    obj1_entry.bind("<FocusOut>", lambda event: on_focusout(obj1_entry, "Primeiro OBJ"))
    obj1_entry.grid(row=2, column=1, padx=(0, 5), pady=5, sticky="ew")

    obj2_entry = tk.Entry(conjunto_frame, font=("Helvetica", 10, "bold"))
    obj2_entry.insert(0, "Segundo OBJ")
    obj2_entry.config(fg="grey")
    obj2_entry.bind("<FocusIn>", lambda event: on_entry_click(obj2_entry, "Segundo OBJ"))
    obj2_entry.bind("<FocusOut>", lambda event: on_focusout(obj2_entry, "Segundo OBJ"))
    obj2_entry.grid(row=2, column=3, padx=(0, 5), pady=5, sticky="ew")

    return objeto1_entry, objeto2_entry, obj1_entry, obj2_entry

def on_entry_click(entry, placeholder):
    if entry.cget("fg") == "grey" and entry.get() == placeholder:
        entry.delete(0, "end")
        entry.config(fg="black")

def on_focusout(entry, placeholder):
    if entry.get() == "":
        entry.insert(0, placeholder)
        entry.config(fg="grey")

def reiniciar_programa():
    try:
        python = sys.executable
        os.chdir(os.path.dirname(sys.argv[0]))
        subprocess.Popen([python] + sys.argv)
        root.destroy()
    except Exception as e:
        print(f"Erro ao reiniciar o programa: {e}")

def mostrar_erro(mensagem):
    global erro_exibido
    if not erro_exibido:
        erro_window = tk.Toplevel()
        erro_window.title("Erro")
        label_erro = tk.Label(erro_window, text=mensagem, font=("Helvetica", 10))
        label_erro.pack(padx=20, pady=10)
        ok_button = tk.Button(erro_window, text="OK", font=("Helvetica", 10, "bold"), command=erro_window.destroy)
        ok_button.pack(pady=5)
        erro_window.geometry("470x100")
        erro_window.resizable(False, False)
        x_erro = (root.winfo_screenwidth() - erro_window.winfo_reqwidth()) / 2
        y_erro = (root.winfo_screenheight() - erro_window.winfo_reqheight()) / 2
        erro_window.geometry("+%d+%d" % (x_erro, y_erro))
        erro_exibido = True

def mostrar_erro2(mensagem):
    # Mesma lógica, ou pode reutilizar mostrar_erro
    mostrar_erro(mensagem) 

def validar_quantidade_entry(*args):
    quantidade_texto = quantidade_entry.get()
    if quantidade_texto == "": return
    if not quantidade_texto.isdigit():
        mostrar_erro("Por favor, digite apenas números inteiros para a quantidade de conjuntos.")
        quantidade_entry.delete(0, tk.END)

def criar_conjuntos():
    global erro_exibido, global_workbook, conjuntos_objetos
    erro_exibido = False
    
    try:
        num_conjuntos = int(quantidade_entry.get())
    except ValueError:
        mostrar_erro2("Por favor, digite um número válido para a quantidade de conjuntos.")
        return

    if num_conjuntos == 0:
        limite_label.config(text="Mínimo de planilhas é 1.", fg="red")
        return
    if num_conjuntos > 100:
        limite_label.config(text="Limite máximo de 100 planilhas alcançado.", fg="red")
        return
    else:
        limite_label.config(text="")

    conjuntos_objetos = []
    for child in conjuntos_frame.winfo_children():
        child.destroy()

    global_workbook = openpyxl.Workbook() # Reseta o workbook aqui

    for i in range(1, num_conjuntos + 1):
        primeiro_objeto_entry, segundo_objeto_entry, primeiro_obj_entry, segundo_obj_entry = criar_conjunto_labels(conjuntos_frame, i)
        conjuntos_objetos.append((primeiro_objeto_entry, segundo_objeto_entry, primeiro_obj_entry, segundo_obj_entry))
    
    # Remove a planilha padrão 'Sheet' se existirem outras, ou deixa para depois
    # Por enquanto, salvamos assim
    # global_workbook.save(global_excel_filename_obj) # Salvar aqui pode ser prematuro, mas estava no original

    conjuntos_canvas.update_idletasks()
    altura_conjuntos = conjuntos_frame.winfo_reqheight()
    conjuntos_canvas.config(scrollregion=(0, 0, conjuntos_canvas.winfo_width(), altura_conjuntos))

    altura_canvas = conjuntos_canvas.winfo_height()
    if num_conjuntos > 0:
        scrollbar.pack(side="right", fill="y")
        altura_scrollbar = min(1.0, altura_canvas / altura_conjuntos)
        scrollbar.set(0, altura_scrollbar)
    else:
        scrollbar.pack_forget()
    liberar_botaoum()

def liberar_criar_conjuntos():
    if caminho_arquivo1 and caminho_entry1.get():
        criar_button.config(state="normal")
    else:
        criar_button.config(state="disabled")
        for child in conjuntos_frame.winfo_children():
            child.destroy()
    root.after(100, liberar_criar_conjuntos)

def liberar_botaoum():
    global conjuntos_objetos, caminho_arquivo1
    all_entries_filled = True
    for objeto1_entry, objeto2_entry, obj1_entry, obj2_entry in conjuntos_objetos:
        if objeto1_entry.winfo_exists():
             if not objeto1_entry.get() or objeto1_entry.get() == "Primeiro par_objeto" \
                or not objeto2_entry.get() or objeto2_entry.get() == "Segundo par_objeto" \
                or not obj1_entry.get() or obj1_entry.get() == "Primeiro OBJ" \
                or not obj2_entry.get() or obj2_entry.get() == "Segundo OBJ":
                all_entries_filled = False
                break
        else:
             all_entries_filled = False # Se não existem, não estão preenchidos

    if all_entries_filled and caminho_arquivo1:
        procurar_button1.config(state="normal")
    else:
        procurar_button1.config(state="disabled")
    root.after(100, liberar_botaoum)

def liberar_botaodois():
    if caminho_arquivo2 and caminho_entry2.get():
        procurar_button2.config(state="normal")
    else:
        procurar_button2.config(state="disabled")
    root.after(100, liberar_botaodois)

def procurar_objetos():
    global caminho_arquivo1, conjuntos_objetos, global_workbook
    
    # Se global_workbook não foi inicializado corretamente ou precisa ser limpo para o arquivo OBJ
    # Nota: No original, 'criar_conjuntos' reseta o workbook. 
    # Aqui vamos usar o que foi criado lá. Removemos 'Sheet' padrão se houver.
    if 'Sheet' in global_workbook.sheetnames:
        del global_workbook['Sheet']

    for primeiro_objeto, segundo_objeto, primeiro_obj, segundo_obj in conjuntos_objetos:
        procurar(primeiro_objeto.get().upper(), segundo_objeto.get().upper(), primeiro_obj.get().upper(),
                 segundo_obj.get().upper(), caminho_arquivo1, global_workbook, colunas_desejadas)
    
    global_workbook.save(global_excel_filename_obj)
    root.withdraw()
    mostrar_alerta()

# --- FUNÇÃO PRINCIPAL MODIFICADA ---
def organizar_distvel():
    global caminho_arquivo2, global_workbook
    
    # Reinicializa o workbook para Dist/Vel para evitar misturar com Obj se o usuário não reiniciar
    global_workbook = openpyxl.Workbook()
    
    # Realiza a organização dos dados (cria as abas 1, 2, 3...)
    organizar(caminho_arquivo2, global_workbook)
    
    # Remove a folha padrão "Sheet" criada pelo openpyxl.Workbook(), se ela estiver vazia/não usada
    # E NÃO remove as abas numéricas criadas
    if 'Sheet' in global_workbook.sheetnames:
        del global_workbook['Sheet']
        
    # Salva o arquivo Excel
    global_workbook.save(global_excel_filename_distvel)
    
    # Mostra a mensagem de alerta
    mostrar_alerta()

def mostrar_alerta():
    alerta = tk.Toplevel()
    alerta.title("Alerta")
    alerta.configure(bg="#f0f0f0")
    alerta.geometry("400x100")
    
    mensagem_label = tk.Label(alerta, text="Dados filtrados com sucesso!\nVerifique o arquivo gerado.", font=("Helvetica", 12), bg="#f0f0f0")
    mensagem_label.pack(padx=20, pady=10)
    
    ok_button = tk.Button(alerta, text="OK", font=("Helvetica", 12, "bold"), bg="#4CAF50", fg="white", command=partial(fechar_alerta, alerta))
    ok_button.pack(pady=5)
    
    alerta.resizable(False, False)
    largura_tela = alerta.winfo_screenwidth()
    altura_tela = alerta.winfo_screenheight()
    x_alerta = (largura_tela - 400) / 2
    y_alerta = (altura_tela - 100) / 2
    alerta.geometry("+%d+%d" % (x_alerta, y_alerta))
    alerta.protocol("WM_DELETE_WINDOW", lambda: None)

def fechar_alerta(alerta):
    alerta.destroy()
    abrir_arquivo_excel()

def abrir_arquivo_excel():
    if os.path.exists(global_excel_filename_obj) and os.path.exists(global_excel_filename_distvel):
        tempo_mod_obj = os.path.getmtime(global_excel_filename_obj)
        tempo_mod_distvel = os.path.getmtime(global_excel_filename_distvel)
        arquivo_recente = global_excel_filename_obj if tempo_mod_obj > tempo_mod_distvel else global_excel_filename_distvel
        os.system(f'start excel "{arquivo_recente}"')
    elif os.path.exists(global_excel_filename_obj):
        os.system(f'start excel "{global_excel_filename_obj}"')
    elif os.path.exists(global_excel_filename_distvel):
        os.system(f'start excel "{global_excel_filename_distvel}"')
    else:
        print("Nenhum arquivo encontrado.")
    root.quit()

def procurar_colunas(caminho_arquivo):
    global pares_objetos, objs
    try:
        df = pd.read_excel(caminho_arquivo, header=6)
        objetos = set(df['OBJECTS'].astype(str))
        events = set(df['Events'].astype(str))

        pares_objetos.clear()
        for objeto in objetos:
            if objeto.strip() and objeto != 'nan':
                pares = objeto.split(' & ')
                pares_objetos.update(pares)
        
        objs.clear()
        for event in events:
            if "OBJ" in event:
                try:
                    obj = event.split("OBJ")[1].split()[0]
                    objs.add(obj)
                except IndexError:
                    pass
        
    except Exception as e:
        print(f"Erro ao abrir o arquivo Excel para leitura de objetos: {e}")

def atualizar_rotulos():
    global pares_objetos, objs
    if caminho_arquivo1:
        procurar_colunas(caminho_arquivo1)
        texto_pares = f"Pares de Objetos: {', '.join(sorted(pares_objetos))}"
        texto_objs = f"OBJs: {', '.join(sorted(objs))}"
        pares_objetos_var.set(texto_pares)
        objs_var.set(texto_objs)

# --- Construção da GUI ---

root = tk.Tk()
root.title("AUTOMATIZADO")
largura_tela = root.winfo_screenwidth()
altura_tela = root.winfo_screenheight()

pesquisa_frame = tk.Frame(root)
pesquisa_frame.grid(row=0, column=0, columnspan=2, pady=10)

pesquisar1_button = tk.Button(pesquisa_frame, text="Pesquisar Arquivo (OBJ)", font=("Helvetica", 10, "bold"), bg="#4CAF50", fg="white", command=pesquisar_arquivo1)
pesquisar1_button.grid(row=0, column=0, padx=10, pady=5)

pesquisar2_button = tk.Button(pesquisa_frame, text="Pesquisar Arquivo (DIST/VEL)", font=("Helvetica", 10, "bold"), bg="#4CAF50", fg="white", command=pesquisar_arquivo2)
pesquisar2_button.grid(row=0, column=1, padx=10, pady=5)

caminho_entry1 = tk.Entry(root, width=50, **entry_label_style)
caminho_entry1.grid(row=1, column=1, padx=(0, 5), pady=5)
limpar1_button = tk.Button(root, text="x", font=("Arial", 10, "bold"), fg="red", bd=0, highlightthickness=0, command=limpar_entry1)
limpar1_button.grid(row=1, column=0, sticky="e", padx=(5, 2), pady=5)

caminho_entry2 = tk.Entry(root, width=50, **entry_label_style)
caminho_entry2.grid(row=2, column=1, padx=(0, 5), pady=5)
limpar2_button = tk.Button(root, text="x", font=("Arial", 10, "bold"), fg="red", bd=0, highlightthickness=0, command=limpar_entry2)
limpar2_button.grid(row=2, column=0, sticky="e", padx=(5, 2), pady=5)

criar_conjuntos_frame = tk.Frame(root)
criar_conjuntos_frame.grid(row=3, column=0, columnspan=2, pady=10)

quantidade_label = tk.Label(criar_conjuntos_frame, text="Quantidade de conjuntos:", font=("Helvetica", 10, "bold"))
quantidade_label.grid(row=0, column=0, padx=5)

quantidade_entry = tk.Entry(criar_conjuntos_frame, width=15, **entry_label_style)
quantidade_entry.grid(row=0, column=1, padx=5)
quantidade_entry.bind("<KeyRelease>", validar_quantidade_entry)

criar_button = tk.Button(criar_conjuntos_frame, text="Criar Conjuntos", font=("Helvetica", 10, "bold"), bg="#4CAF50", fg="white", command=criar_conjuntos, state="disabled")
criar_button.grid(row=0, column=2, padx=5)

atualizar_rotulos_button = tk.Button(criar_conjuntos_frame, text="Ver objetos", font=("Helvetica", 10, "bold"), bg="#4CAF50", fg="white", command=atualizar_rotulos)
atualizar_rotulos_button.grid(row=1, column=2, padx=5, pady=5)

pares_objetos_var = tk.StringVar()
pares_objetos_value_label = tk.Label(criar_conjuntos_frame, textvariable=pares_objetos_var, font=("Helvetica", 10), wraplength=500, justify="left")
pares_objetos_value_label.grid(row=3, column=0, columnspan=3, padx=10, pady=(0, 5), sticky="w")

objs_var = tk.StringVar()
objs_value_label = tk.Label(criar_conjuntos_frame, textvariable=objs_var, font=("Helvetica", 10), wraplength=500, justify="left")
objs_value_label.grid(row=4, column=1, columnspan=2, pady=(0, 5), sticky="w")

limite_label = tk.Label(criar_conjuntos_frame, text="", fg="red")
limite_label.grid(row=1, columnspan=3, pady=(5, 0))

conjuntos_scroll_frame = tk.Frame(root)
conjuntos_scroll_frame.grid(row=4, column=0, columnspan=2, pady=10)

conjuntos_canvas = tk.Canvas(conjuntos_scroll_frame, bg="#E0E0E0")
conjuntos_canvas.pack(side="left", fill="both", expand=True)

scrollbar = tk.Scrollbar(conjuntos_scroll_frame, orient="vertical", command=conjuntos_canvas.yview)
scrollbar.pack(side="right", fill="y")

conjuntos_frame = tk.Frame(conjuntos_canvas, bg="#E0E0E0")
conjuntos_frame.grid(row=0, column=0, sticky="nsew")
conjuntos_canvas.create_window((0, 0), window=conjuntos_frame, anchor="nw")

procurar_button_frame = tk.Frame(root)
procurar_button_frame.grid(row=5, column=0, columnspan=2, pady=10)

procurar_button1 = tk.Button(procurar_button_frame, text="Procurar Objetos", font=("Helvetica", 12, "bold"), bg="#4CAF50", fg="white", command=procurar_objetos, state="disabled")
procurar_button1.pack(side="left", padx=10)

procurar_button2 = tk.Button(procurar_button_frame, text="Organizar Dist/Vel", font=("Helvetica", 12, "bold"), bg="#4CAF50", fg="white", command=organizar_distvel, state="disabled")
procurar_button2.pack(side="left", padx=10)

reiniciar_button = tk.Button(procurar_button_frame, text="Reiniciar", font=("Helvetica", 10, "bold"), bg="#4CAF50", fg="white", command=reiniciar_programa)
reiniciar_button.pack(side="left", padx=10)

root.resizable(False, False)
root.update_idletasks()
x_root = (largura_tela - root.winfo_width()) / 2
y_root = (altura_tela - root.winfo_height()) / 2
root.geometry("+%d+%d" % (x_root, y_root))

try:
    caminho_icone = "memorylab.ico"
    if os.path.exists(caminho_icone):
        icone = Image.open(caminho_icone)
        icone_tk = ImageTk.PhotoImage(icone)
        root.iconphoto(True, icone_tk)
except Exception:
    pass

root.mainloop()
