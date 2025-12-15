import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

def organizar(caminho_arquivo2, global_workbook):
    # Carregar dados do Excel
    try:
        df = pd.read_excel(caminho_arquivo2, header=6, sheet_name="Bin Measure")
    except Exception as e:
        print(f"Erro ao ler a aba 'Bin Measure': {e}")
        return []

    # --- TRATAMENTO DOS DIAS ---
    # Converta os valores na coluna 'DAY' para strings
    df['DAY'] = df['DAY'].astype(str)
    # Remova o ponto decimal e zeros à direita ('.0')
    df['DAY'] = df['DAY'].str.replace('.0', '', regex=False)
    # Converta para inteiros
    df['DAY'] = pd.to_numeric(df['DAY'], errors='coerce')
    # Remove linhas inválidas
    df = df.dropna(subset=['DAY'])
    
    # --- CÁLCULO DA SOMA (BIN1 + BIN2) ---
    # Garante que Bin1 e Bin2 sejam numéricos, transformando erros em NaN e depois em 0 para a soma
    # Se preferir que "Vazio" + 10 seja NaN, remova o .fillna(0)
    df['Bin1'] = pd.to_numeric(df['Bin1'], errors='coerce').fillna(0)
    
    # Verifica se a coluna Bin2 existe, caso o arquivo mude no futuro
    if 'Bin2' in df.columns:
        df['Bin2'] = pd.to_numeric(df['Bin2'], errors='coerce').fillna(0)
        # Cria a coluna da soma
        df['Bin_Soma'] = df['Bin1'] + df['Bin2']
    else:
        # Se não tiver Bin2, a soma é igual ao Bin1
        df['Bin_Soma'] = df['Bin1']

    # Ordena os dias para criar as abas em ordem
    dias_unicos = sorted(df['DAY'].unique().astype(int))

    # Definições de Eventos
    coluna_evento = 'Measure'
    evento_distance_traveled = 'Mouse 1 Center Distance Traveled (apart 1.000000 second)'
    evento_velocity_average = 'Mouse 1 Center Velocity Average (apart 1.000000 second)'

    # Dicionário para renomear colunas (Note que agora usamos 'Bin_Soma')
    novo_nome_colunas = {
        'DAY': 'Dia',
        'ANIMAL': 'Animal',
        'Bin_Soma_dist': 'Distância',       # Alterado de Bin1 para Bin_Soma
        'Bin_Soma_dist_blank': 'Distância (metros)',
        'Bin_Soma_vel': 'Velocidade',      # Alterado de Bin1 para Bin_Soma
    }

    # Itera sobre cada dia encontrado no arquivo
    for dia in dias_unicos:
        nome_aba = str(dia)
        
        # Cria ou seleciona a aba
        if nome_aba in global_workbook.sheetnames:
            ws = global_workbook[nome_aba]
        else:
            ws = global_workbook.create_sheet(title=nome_aba)

        # Filtra os dados usando a coluna já calculada (Bin_Soma)
        dados_distance = df[(df[coluna_evento] == evento_distance_traveled) & (df['DAY'] == dia)]
        dados_velocity = df[(df[coluna_evento] == evento_velocity_average) & (df['DAY'] == dia)]

        if dados_distance.empty and dados_velocity.empty:
            continue

        # Merge dos dados usando 'DAY' e 'ANIMAL' como chave
        # Isso vai criar colunas: Bin_Soma_dist e Bin_Soma_vel
        df_dia = pd.merge(dados_distance, dados_velocity, on=['DAY', 'ANIMAL'], suffixes=('_dist', '_vel'))

        # Seleciona apenas as colunas de interesse (Soma)
        colunas_para_manter = ['DAY', 'ANIMAL', 'Bin_Soma_dist', 'Bin_Soma_vel']
        
        # Garante que as colunas existem antes de reindexar
        colunas_existentes = [c for c in colunas_para_manter if c in df_dia.columns]
        df_dia = df_dia.reindex(columns=colunas_existentes)

        # Adiciona a coluna de Distância em Metros
        if 'Bin_Soma_dist' in df_dia.columns:
            # Insere a coluna em branco logo após a distância
            posicao_dist = df_dia.columns.get_loc('Bin_Soma_dist') + 1
            df_dia.insert(posicao_dist, 'Bin_Soma_dist_blank', '')
            
            # Preenche dividindo por 1000
            df_dia['Bin_Soma_dist_blank'] = df_dia['Bin_Soma_dist'] / 1000

        # Adiciona os dados à aba
        for row in dataframe_to_rows(df_dia, index=False, header=True):
            ws.append(row)

        # --- FORMATAÇÃO VISUAL ---
        
        # Ajuste de larguras
        colunas_menores = ['C', 'E'] # C=Distância, E=Velocidade
        coluna_maior = ['D']         # D=Metros

        for letra in colunas_menores:
            ws.column_dimensions[letra].width = 14
        for letra in coluna_maior:
            ws.column_dimensions[letra].width = 18

        # Renomear Cabeçalhos (Aplica o dicionário novo_nome_colunas)
        for cell in ws[1]:
            if cell.value in novo_nome_colunas:
                cell.value = novo_nome_colunas[cell.value]

        # Estilização
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                              top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.border = border_style

    return dias_unicos
